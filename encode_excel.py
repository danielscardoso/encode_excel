import openpyxl
import chardet
import os
from urllib.parse import unquote
from difflib import SequenceMatcher
from openpyxl.styles import PatternFill

# Code created by Daniel Cardoso - danielscardoso@gmail.com - on January 2025, with the purpose of correcting a specific encoding issue in an Excel file. 
# This code is aimed at a specific use case and may not work as intended for other scenarios. Use at your own risk. 
# It was made with European Portuguese in mind as its output scenario.
# It was produced with the help of Copilot on Microsoft's Visual Studio Code.
# Creative Commons BY-NC-SA 4.0 https://creativecommons.org/licenses/by-nc-sa/4.0/ 

# Dictionary of common European Portuguese nouns
portuguese_nouns = {
    "p√≠lula": "pílula",
    "√©": "é",
    "Ãª": "ê",
    "√¢": "â",
    "√§": "ç",
    "√°": "á",
    "√º": "ú",
    "√£": "ã",
    "√´": "ô",
    "√±": "ñ",
    "√≥": "ó",
    "√ß": "ç",
    "√µ": "õ",
    "Ã³": "ó",
    "viol√™ncia": "violência",
    "Ap√≠lula": "Apílula",
    # Add more common nouns as needed
}

def detect_and_convert_encoding(text):
    """
    Detects the character encoding of the given text and converts it to Unicode if necessary.

    Args:
        text: The text string to be analyzed.

    Returns:
        The text string encoded in Unicode.
    """
    if isinstance(text, str):
        # Detect encoding
        byte_text = text.encode('raw_unicode_escape', errors='ignore')  # Use 'raw_unicode_escape' to ensure all characters are encoded
        result = chardet.detect(byte_text)
        encoding = result['encoding']
        confidence = result['confidence']
        
        if encoding and confidence > 0.5:
            try:
                text = byte_text.decode(encoding)
            except (UnicodeDecodeError, TypeError):
                pass
        else:
            # Try common encodings if detection fails
            for enc in ['utf-8', 'latin1', 'iso-8859-1']:
                try:
                    text = byte_text.decode(enc)
                    break
                except (UnicodeDecodeError, TypeError):
                    pass
            else:
                # If all else fails, decode using 'utf-8' with replacement for errors
                text = byte_text.decode('utf-8', errors='replace')
        
        # Attempt to replace errors using the dictionary of common nouns
        for key, value in portuguese_nouns.items():
            text = text.replace(key, value)
    return text

def decode_url(text):
    """
    Decodes a URL-encoded string.

    Args:
        text: The URL-encoded string to be decoded.

    Returns:
        The decoded URL string.
    """
    return unquote(text)

def find_similar_url(url, decoded_urls):
    """
    Finds a similar URL in the decoded_urls dictionary.

    Args:
        url: The URL string to be analyzed.
        decoded_urls: A dictionary of previously decoded URLs.

    Returns:
        The decoded URL string if a similar URL is found, otherwise None.
    """
    for decoded_url in decoded_urls:
        similarity = SequenceMatcher(None, url, decoded_url).ratio()
        if similarity > 0.9:
            return decoded_urls[decoded_url]
    return None

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the file path
file_path = os.path.join(script_dir, 'teste.xlsx')

decoded_urls = {}

# Define the purple fill for cells with encoding errors
purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")

try:
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    
    # Iterate through all sheets in the workbook
    for sheet in workbook.worksheets:
        # Iterate through all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Check for similar URL in the decoded_urls dictionary
                    similar_url = find_similar_url(cell.value, decoded_urls)
                    if similar_url:
                        cell.value = similar_url
                    else:
                        # First, detect and convert encoding
                        new_value = detect_and_convert_encoding(cell.value)
                        # Then, decode the URL
                        decoded_value = decode_url(new_value)
                        if decoded_value != cell.value:
                            cell.value = decoded_value
                            decoded_urls[cell.value] = decoded_value
                            # Color the cell purple if there was an encoding error
                            if '�' in decoded_value:
                                cell.fill = purple_fill

    # Save the modified workbook with a new name
    new_file_path = os.path.join(script_dir, 'teste_unicode_allpages.xlsx')
    workbook.save(new_file_path)
    print(f"Workbook saved as '{new_file_path}'")
except FileNotFoundError:
    print(f"Error: The file '{file_path}' was not found.")